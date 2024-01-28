from datetime import datetime
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone

from rest_framework import status
from rest_framework.response import Response

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook import Workbook

desired_column_order = [
    'id', 'name', 'surname', 'email', 'phone', 'age', 'course', 'course_format','course_type', 'already_paid',
    'sum', 'created_at', 'status', 'group_id', 'manager_id']

column_widths = {
    'id': 5,
    'name': 15,
    'surname': 15,
    'email': 36,
    'phone': 20,
    'age': 5,
    'course': 10,
    'course_format': 15,
    'course_type': 15,
    'already_paid': 15,
    'sum': 10,
    'created_at': 20,
    'status': 12,
    'group_id': 10,
    'manager_id': 12
}


def __table_style(df, sheet):
    __create_header(sheet)
    for row_index, row in enumerate(df.itertuples(), start=2):
        if row_index % 2 == 0:
            for col_index in range(1, len(desired_column_order) + 1):
                cell = sheet.cell(row=row_index, column=col_index)
                cell.fill = PatternFill(start_color='E9E9E9', end_color='E9E9E9', fill_type='solid')
        for col_index, column_name in enumerate(desired_column_order, start=1):
            value = getattr(row, column_name)
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            width = column_widths.get(column_name, 12)
            sheet.column_dimensions[sheet.cell(row=1, column=col_index).column_letter].width = width


def __create_header(sheet):
    for c_idx, column_name in enumerate(desired_column_order, start=1):
        cell = sheet.cell(row=1, column=c_idx, value=column_name)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='50C878', end_color='50C878', fill_type='solid')
        cell.font = Font(color='FFFFFF')


def date_converter(data):
    for item in data:
        for key, value in item.items():
            if isinstance(value, datetime):
                item[key] = timezone.localtime(value).replace(tzinfo=None)


def name_creator():
    current_datetime = datetime.now()
    formatted_datetime = current_datetime.strftime('%Y-%m-%d')
    return f'{formatted_datetime}.xlsx'


def book_creator(df, filename):
    try:
        workbook = Workbook()
        sheet = workbook.active
        excel_buffer = BytesIO()
        __table_style(df, sheet)
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        excel_buffer_content = excel_buffer.getvalue()
        response = HttpResponse(
            excel_buffer_content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
