from datetime import datetime
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone

from rest_framework import status
from rest_framework.response import Response

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook import Workbook

from configs.extra_conf import column_widths, desired_column_order

from core.models import ProfileModel

from apps.groups.models import GroupModel


class ExportFileService:

    @staticmethod
    def name_creator():
        current_datetime = datetime.now()
        formatted_datetime = current_datetime.strftime('%Y-%m-%d')
        return f'{formatted_datetime}.xlsx'

    @classmethod
    def table_creator(cls, orders):
        data = list(orders.values())
        cls.__data_converter(data)
        return pd.DataFrame(data)

    @classmethod
    def book_creator(cls, df, filename):
        try:
            workbook = Workbook()
            sheet = workbook.active
            excel_buffer = BytesIO()
            cls.__table_style(df, sheet)
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            excel_buffer_content = excel_buffer.getvalue()
            response = HttpResponse(
                excel_buffer_content,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename={filename}'
            return response
        except Exception as e:
            return Response({'error exel book creator': str(e)}, status.HTTP_503_SERVICE_UNAVAILABLE)

    @staticmethod
    def __data_converter(data):
        group_id = set(item['group_id'] for item in data)
        manager_id = set(item['manager_id'] for item in data)
        groups = {group.id: group.name for group in GroupModel.objects.filter(pk__in=group_id)}
        managers = {manager.id: manager.name for manager in ProfileModel.objects.filter(pk__in=manager_id)}
        for item in data:
            if isinstance(item['created_at'], datetime):
                item['created_at'] = timezone.localtime(item['created_at']).replace(tzinfo=None)
            item['group_id'] = groups.get(item['group_id'], '')
            item['manager_id'] = managers.get(item['manager_id'], '')

    @staticmethod
    def __create_header(sheet):
        replace_column_name = {
            'group_id': 'group',
            'manager_id': 'manager'
        }
        for column_index, column_name in enumerate(desired_column_order, start=1):
            column_name_table = replace_column_name.get(column_name, column_name)
            cell = sheet.cell(row=1, column=column_index, value=column_name_table)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='50C878', end_color='50C878', fill_type='solid')
            cell.font = Font(color='FFFFFF')

    @classmethod
    def __table_style(cls, df, sheet):
        cls.__create_header(sheet)
        for row_index, row in enumerate(df.itertuples(), start=2):
            if row_index % 2 == 0:
                for column_index in range(1, len(desired_column_order) + 1):
                    cell = sheet.cell(row=row_index, column=column_index)
                    cell.fill = PatternFill(start_color='E9E9E9', end_color='E9E9E9', fill_type='solid')
            for column_index, column_name in enumerate(desired_column_order, start=1):
                value = getattr(row, column_name)
                cell = sheet.cell(row=row_index, column=column_index, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                width = column_widths.get(column_name, 12)
                sheet.column_dimensions[sheet.cell(row=1, column=column_index).column_letter].width = width
